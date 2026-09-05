"""Verificaciones financieras: python -m unittest discover -s finanzas -p 'test_*.py'."""
import copy
import unittest
from pathlib import Path
import openpyxl

try:
    from . import modelo as m
except ImportError:
    import modelo as m


class FinanceTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.base=m.funding()

    def test_analytic_npv_irr_and_payback(self):
        flows=[-100]+[0]*11+[121]
        self.assertAlmostEqual(m.irr_monthly(flows),(1.21)**(1/12)-1,places=10)
        self.assertAlmostEqual(sum(f/1.1**(i/12) for i,f in enumerate(flows)),10)
        self.assertAlmostEqual(m.payback_months([-100,40,40,40]),2.5)
        self.assertAlmostEqual(m.payback_months(flows,.21),12)
        self.assertIsNone(m.payback_months([-100,10,20]))

    def test_negative_and_multiple_irr(self):
        self.assertAlmostEqual(m.irr_monthly([-100,90]),-.1)
        self.assertIsNone(m.irr_monthly([-100,230,-132]))  # 10% y 20% mensuales
        self.assertIsNone(m.irr_monthly([-100,-10]))

    def test_project_independent_of_debt(self):
        a=m.project_flows(m.simulate(principal=0))
        b=m.project_flows(m.simulate(principal=40000000,term=5))
        for x,y in zip(a,b): self.assertAlmostEqual(x,y,places=6)
        # La inversión y variaciones de caja operativa se comprometen una sola vez.
        r=self.base
        terminal_reserve=r['reserve_months']*r['rows'][59]['fijos']+r['contingency']
        self.assertAlmostEqual(sum(m.project_flows(r)),
            -r['initial_uses']+sum(row['fcff'] for row in r['rows'][:60])-terminal_reserve,places=6)

    def test_scenario_irr_residual(self):
        for factor,mf in ((1,1),(.65,1),(1.5,1.5)):
            r=m.simulate(m.S,self.base['principal'],10,factor,mf)
            k=m.indicators(r,.2)
            self.assertIsNotNone(k['tir_anual'])
            value=sum(f/(1+k['tir_anual'])**(i/12) for i,f in enumerate(m.project_flows(r)))
            self.assertAlmostEqual(value,0,delta=.01)
            if k['payback'] is not None:
                flows=m.project_flows(r)
                month=int(k['payback'])
                self.assertLess(sum(flows[:month+1]),0)
                self.assertGreaterEqual(sum(flows[:month+2]),0)

    def test_cash_inventory_and_debt_reconcile(self):
        r=self.base
        cash=r['initial_cash'];stock=r['initial_inventory'];inventory=stock*r['bom'];debt=r['principal']
        for row in r['rows']:
            cash+=row['fcfe'];stock+=row['compras_unidades']-row['unidades']
            inventory+=row['compras_netas']-row['costo_vendido'];debt-=row['amortizacion']
            self.assertAlmostEqual(cash,row['caja'],places=6)
            self.assertEqual(stock,row['stock_unidades'])
            self.assertGreaterEqual(stock,0)
            self.assertAlmostEqual(inventory,row['inventario_neto'],places=6)
            self.assertAlmostEqual(debt,row['saldo_deuda'],places=5)
            self.assertAlmostEqual(row['fcfe'],row['fcff']+row['impuesto_proyecto']-row['impuesto_reservado']-row['servicio_deuda'],places=6)
        self.assertAlmostEqual(debt,0,places=5)
        self.assertAlmostEqual(m.S['aporte_socios']+r['principal'],r['initial_uses']+r['fees']+r['initial_cash'])

    def test_initial_vat_credit(self):
        r=self.base
        expected=r['initial_inventory']*(r['bom']-900)*m.S['iva']
        self.assertAlmostEqual(expected,r['initial_vat'])
        first=r['rows'][0]
        purchases=(first['compras_netas']-900*first['compras_unidades'])*m.S['iva']
        used=min(first['ventas_brutas']-first['ventas_netas'],expected)
        self.assertAlmostEqual(first['credito_iva'],expected+purchases-used)
        self.assertAlmostEqual(first['iva_inmovilizado'],purchases-used)

    def test_bme_cost_propagates(self):
        s=copy.deepcopy(m.S)
        line=next(x for x in s['bom'] if 'BME280 I2C' in x[0])
        self.assertAlmostEqual(line[2]*1.19,3500)
        line[2]+=1000
        r=m.simulate(s,self.base['principal'])
        self.assertAlmostEqual(r['bom']-self.base['bom'],1000)
        self.assertAlmostEqual(r['initial_uses']-self.base['initial_uses'],self.base['initial_inventory']*1190)
        self.assertLess(m.yearly(r)[0]['ebitda'],m.yearly(self.base)[0]['ebitda'])
        self.assertLess(m.indicators(r,.2)['van'],m.indicators(self.base,.2)['van'])
        self.assertEqual(s['api_clima_mensual'],0)
        self.assertTrue(any('Carcasa impresa en 3D' in x[0] for x in s['bom']))

    def test_funding_is_minimum_and_reserve_stops_debt(self):
        self.assertGreaterEqual(min(x['caja_libre'] for x in self.base['rows'][:24]),0)
        smaller=m.simulate(m.S,self.base['principal']-100000)
        self.assertTrue(smaller['initial_cash']<smaller['rows'][0]['reserva'] or min(x['caja_libre'] for x in smaller['rows'][:24])<0)
        row=self.base['rows'][120]
        self.assertAlmostEqual(row['reserva'],3*row['fijos']+self.base['contingency'])

    def test_zero_sales(self):
        r=m.simulate(factor=0)
        self.assertIsNone(m.yearly(r)[0]['equilibrio_operativo'])
        self.assertIsNone(m.indicators(r,.2)['payback'])

    def test_generated_workbook_matches_model(self):
        path=m.ROOT/'Flujo de caja y financiamiento - TerraSense.xlsx'
        wb=openpyxl.load_workbook(path,data_only=True,read_only=True)
        try:
            k=m.indicators(self.base,.2)
            values=list(wb['Evaluacion proyecto'].values)[1]
            for actual,key in zip(values[1:6],('inversion','van','tir_anual','payback','payback_descontado')):
                self.assertAlmostEqual(actual,k[key],places=5)
            for actual,expected in zip(list(wb['Base_mensual'].values)[1:],self.base['rows']):
                for value,target in zip(actual,expected.values()):
                    self.assertAlmostEqual(value,target,delta=.0001)
            for sheet in wb:
                for row in sheet:
                    for cell in row:
                        self.assertNotEqual(cell.data_type,'e',f'{sheet.title}!{getattr(cell,"coordinate", "vacía")}')
        finally:
            wb.close()
        bom=openpyxl.load_workbook(m.ROOT/'PCB/BOM_TerraSense.xlsx',data_only=True,read_only=True)
        try:
            self.assertAlmostEqual(bom.active.cell(len(m.S['bom'])+2,4).value,self.base['bom'])
        finally:
            bom.close()

    def test_report_financial_blocks_match_model(self):
        try:
            from .informe import REPORT, report_blocks
        except ImportError:
            from informe import REPORT, report_blocks
        variants={'Base':self.base,
                  'Estres':m.simulate(m.S,self.base['principal'],10,m.S['ventas_estres_factor']),
                  'Crecimiento':m.simulate(m.S,self.base['principal'],10,m.S['ventas_crecimiento_factor'],m.S['ventas_crecimiento_factor'])}
        text=REPORT.read_text(encoding='utf-8')
        for name,expected in report_blocks(self.base,variants,m).items():
            actual=text.split(f'<!-- INFORME:{name}:INICIO -->')[1].split(f'<!-- INFORME:{name}:FIN -->')[0].strip()
            self.assertEqual(actual,expected)

    def test_report_energy_example_units(self):
        cycle=(40*12+95*3+60)/3600
        self.assertAlmostEqual(cycle,.22916666666666666)
        remaining=24-(10*16+5*60)/3600
        for standby,expected_daily,expected_days in ((.1,7.51,213.0),(1,29.00,55.2),(5,124.49,12.9)):
            daily=cycle*10+18*5/60+standby*remaining+2000*.02/30
            self.assertAlmostEqual(daily,expected_daily,delta=.005)
            self.assertAlmostEqual(1600/daily,expected_days,delta=.05)


if __name__=='__main__':
    unittest.main()
